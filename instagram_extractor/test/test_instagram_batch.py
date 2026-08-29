import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from run_instagram_batch import _excel_path, _retry_result, _write_excel


class TestInstagramBatch(unittest.TestCase):
    def test_failed_retry_preserves_a_previous_non_error_record(self):
        existing = {"status": "ambiguous", "candidates": [{"username": "lead"}]}
        attempted = {"status": "error", "error": "temporary failure"}
        retained = _retry_result(existing, attempted)
        self.assertEqual(retained["status"], "ambiguous")
        self.assertEqual(retained["last_retry_error"], "temporary failure")

    def test_successful_retry_replaces_the_previous_record(self):
        existing = {"status": "ambiguous"}
        attempted = {"status": "matched", "instagram_username": "best"}
        self.assertIs(_retry_result(existing, attempted), attempted)

    def test_failed_retry_does_not_preserve_an_about_identity(self):
        existing = {
            "status": "ambiguous",
            "linkedin_extracted": {"name": "About"},
        }
        attempted = {"status": "error", "error": "invalid top card"}
        self.assertIs(_retry_result(existing, attempted), attempted)

    def test_excel_path_defaults_to_json_basename(self):
        self.assertEqual(
            _excel_path("results/instagram_profiles.json"),
            "results/instagram_profiles.xlsx",
        )

    def test_excel_export_separates_confirmed_and_potential_accounts(self):
        records = {
            "linkedin/confirmed": {
                "linkedin_url": "linkedin/confirmed",
                "status": "matched",
                "instagram_username": "confirmed_account",
                "instagram_url": "https://instagram.com/confirmed_account/",
                "score": 0.91,
                "linkedin_extracted": {"name": "Confirmed Person"},
            },
            "linkedin/potential": {
                "linkedin_url": "linkedin/potential",
                "status": "ambiguous",
                "potential_instagram_username": "potential_account",
                "potential_instagram_url": "https://instagram.com/potential_account/",
                "potential_score": 0.63,
                "twitter_url": "https://x.com/potential_account/",
                "facebook_url": "https://www.facebook.com/potential.account/",
                "linkedin_extracted": {"name": "Potential Person"},
            },
        }
        with TemporaryDirectory() as directory:
            path = Path(directory) / "accounts.xlsx"
            _write_excel(path, list(records), records)
            workbook = load_workbook(path)

        self.assertEqual(workbook.sheetnames, ["Final Accounts", "All Results"])
        self.assertEqual(workbook["Final Accounts"].max_row, 2)
        self.assertEqual(workbook["Final Accounts"]["D2"].value, "confirmed_account")
        self.assertEqual(workbook["All Results"].max_row, 3)
        self.assertEqual(workbook["All Results"]["D3"].value, "potential_account")
        self.assertEqual(workbook["All Results"]["F3"].value, "potential")
        self.assertEqual(workbook["All Results"]["G3"].value, "https://x.com/potential_account/")
        self.assertEqual(
            workbook["All Results"]["H3"].value,
            "https://www.facebook.com/potential.account/",
        )


if __name__ == "__main__":
    unittest.main()
